# -*- coding: utf-8 -*-
"""
order_core.py
部品発注リスト自動補完システム ― 中核ロジック

役割:
  1. Google スプレッドシート「発注履歴リスト」からデータを取得
  2. Excel「新規発注リスト」の型式をキーに、品名/発注先/単価を自動補完
  3. 合計金額(単価×数量)を計算し、未ヒット型式には「新規」フラグを立てる
  4. 補完済み Excel を出力

このモジュールは UI(app.py) からも CLI からも再利用できるよう、
画面表示には依存しない純粋な関数として実装しています。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

# gspread / google-auth は Google 接続時のみ必要なため、関数内で遅延 import する。
# これにより Excel 処理やテンプレート生成は Google ライブラリ無しでも動作する。
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Google API のアクセス範囲。読み取り専用にしておくことで安全性を高める。
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

# 新規発注リストの列構成(1始まり)。ヘッダー名でも検出するが、見つからなければこの位置を使う。
COL_MODEL = 1    # A: 型式      (手入力キー)
COL_MAKER = 2    # B: メーカー  (自動: 型式→メーカー対応表より)
COL_QTY = 3      # C: 数量      (手入力)
COL_PRICE = 4    # D: 単価      (自動: 履歴より)
COL_VENDOR = 5   # E: 発注先    (自動: 履歴=見積/発注元より)
COL_TOTAL = 6    # F: 合計金額  (単価×数量)
COL_SPECURL = 7  # G: 仕様URL   (自動: 型式の検索リンク)

# 列名の表記ゆれを吸収するための候補
HEADER_ALIASES = {
    "model": ["型式", "型番", "品番", "model", "型式番号"],
    "maker": ["メーカー", "maker", "製造元", "製造メーカー", "ブランド", "brand"],
    "name": ["品名", "名称", "品目", "name"],
    "vendor": ["発注先", "仕入先", "購入先", "vendor", "supplier"],
    "price": ["単価", "価格", "unit price", "price"],
    "qty": ["数量", "個数", "qty", "quantity"],
    "total": ["合計金額", "合計", "金額合計", "total"],
    "spec_url": ["仕様url", "仕様書url", "url", "仕様", "リンク", "spec", "specurl"],
    "status": ["状態", "ステータス", "フラグ", "status"],
    "note": ["備考", "メモ", "摘要", "note", "remarks"],
}


# ----------------------------------------------------------------------------
# データ構造
# ----------------------------------------------------------------------------
@dataclass
class HistoryRecord:
    """発注履歴1件分"""
    name: str = ""
    vendor: str = ""
    price: Optional[float] = None
    note: str = ""
    source: str = ""  # 由来(ファイル名など)。フォルダ走査時の追跡用。


@dataclass
class ProcessResult:
    """処理結果のサマリ"""
    total_rows: int = 0          # 型式が入っていた行数
    matched: int = 0             # 履歴にヒットした行数
    new_items: int = 0           # 未ヒット(新規)の行数
    new_models: list = field(default_factory=list)  # 新規だった型式の一覧
    preview: list = field(default_factory=list)      # 画面表示用の行データ


# ----------------------------------------------------------------------------
# 認証 & 履歴取得
# ----------------------------------------------------------------------------
def get_client(credentials_path: str = "credentials.json"):
    """サービスアカウントの鍵ファイルから gspread クライアントを生成する。"""
    import gspread
    from google.oauth2.service_account import Credentials
    creds = Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
    return gspread.authorize(creds)


def _open_spreadsheet(gc, *, url: str = "", key: str = "", name: str = ""):
    """URL / キー / 名前 のいずれかでスプレッドシートを開く。"""
    if url.strip():
        return gc.open_by_url(url.strip())
    if key.strip():
        return gc.open_by_key(key.strip())
    if name.strip():
        return gc.open(name.strip())
    raise ValueError("スプレッドシートの URL・キー・名前 のいずれかを指定してください。")


def _parse_price(value) -> Optional[float]:
    """'¥1,200' や '1200円' のような文字列から数値を取り出す。空なら None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r"[^\d.\-]", "", s)  # 数字・小数点・マイナス以外を除去
    if s in ("", "-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _normalize_model(value) -> str:
    """型式キーを正規化(前後空白除去・内部の改行/連続空白を1スペースに)。"""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def load_history(
    gc,
    *,
    url: str = "",
    key: str = "",
    name: str = "",
    worksheet: str = "",
) -> dict[str, HistoryRecord]:
    """
    発注履歴スプレッドシートを読み込み、型式をキーにした辞書を返す。
    列構成: A型式 B品名 C発注先 D単価 E備考
    1行目が見出し(型式…)の場合は自動でスキップする。
    """
    sh = _open_spreadsheet(gc, url=url, key=key, name=name)
    ws = sh.worksheet(worksheet.strip()) if worksheet.strip() else sh.sheet1
    return _build_history_from_rows(ws.get_all_values())


def _norm_header(text) -> str:
    """見出し比較用の正規化: 空白(半角/全角)除去・小文字化。
    例: '単 価' -> '単価', '品 名 ・ 形 状' -> '品名・形状'
    """
    return re.sub(r"[\s　]+", "", str(text)).lower()


def _find_header_indices(header_row) -> dict[str, int]:
    """
    見出し行(list)から各項目の列インデックス(0始まり)を別名表で判定する。
    スペースを無視し、別名がセル内に含まれていれば一致とみなす(部分一致)。
    例: 見出し '品名・形状・寸法・仕様等' は別名 '品名' を含むので name 扱い。
    """
    idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        text = _norm_header(cell)
        if not text:
            continue
        for key, aliases in HEADER_ALIASES.items():
            if any(_norm_header(a) in text for a in aliases):
                idx.setdefault(key, i)
    return idx


def _build_history_from_rows(rows) -> dict[str, HistoryRecord]:
    """
    行リスト(list[list])から型式キーの履歴辞書を作る。

    1行目が見出し行の場合は、見出し名(型式/品名/発注先(メーカー)/単価/備考)から
    各列の位置を自動判別する。これにより列順が想定と違っても正しく読み取れる。
    見出しが無ければ A型式 B品名 C発注先 D単価 E備考 の固定位置を使う。
    """
    history: dict[str, HistoryRecord] = {}
    if not rows:
        return history

    header_idx = _find_header_indices(rows[0])
    if "model" in header_idx:
        # 見出し行あり: 名前で列を特定(見つからない項目は既定位置にフォールバック)
        col = {
            "model": header_idx.get("model", 0),
            "name": header_idx.get("name", 1),
            "vendor": header_idx.get("vendor", 2),
            "price": header_idx.get("price", 3),
            "note": header_idx.get("note", 4),
        }
        data_rows = rows[1:]
    else:
        # 見出し無し: 固定位置 A-E
        col = {"model": 0, "name": 1, "vendor": 2, "price": 3, "note": 4}
        data_rows = rows

    def cell(row, i):
        return row[i] if 0 <= i < len(row) else ""

    for row in data_rows:
        model = _normalize_model(cell(row, col["model"]))
        if not model:
            continue
        history[model] = HistoryRecord(
            name=str(cell(row, col["name"])).strip(),
            vendor=str(cell(row, col["vendor"])).strip(),
            price=_parse_price(cell(row, col["price"])),
            note=str(cell(row, col["note"])).strip(),
        )
    return history


# ----------------------------------------------------------------------------
# Google Cloud を使わない取得 (方式A: CSV公開URL / 方式B: ローカルファイル)
# ----------------------------------------------------------------------------
def extract_sheet_id(url: str) -> str:
    """ブラウザのスプレッドシートURLからシートIDを取り出す。"""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9\-_]+)", url)
    if not m:
        raise ValueError("スプレッドシートのURLからIDを取得できませんでした。URLを確認してください。")
    return m.group(1)


def extract_gid(url: str) -> Optional[str]:
    """URLからタブのgid(数値)を取り出す。見つからなければ None(先頭タブ)。"""
    m = re.search(r"[#&?]gid=([0-9]+)", url)
    return m.group(1) if m else None


def build_csv_export_url(url: str, gid: Optional[str] = None) -> str:
    """ブラウザURLを、認証不要のCSVエクスポートURLに変換する。"""
    sheet_id = extract_sheet_id(url)
    if gid is None:
        gid = extract_gid(url)
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    if gid is not None:
        csv_url += f"&gid={gid}"
    return csv_url


def load_history_from_csv_url(url: str, gid: Optional[str] = None) -> dict[str, HistoryRecord]:
    """
    公開(リンク共有)されたGoogleスプレッドシートのCSVエクスポートURLから履歴を取得する。
    Google Cloud / 認証情報は不要。シートが「リンクを知っている全員が閲覧可」であること。
    """
    import csv
    import urllib.request

    csv_url = build_csv_export_url(url, gid)
    req = urllib.request.Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        final_url = resp.geturl()
        raw = resp.read()

    # 非公開シートだとGoogleはHTMLのログイン画面を返す。CSVでないと判断したら明示エラー。
    if "accounts.google.com" in final_url or raw[:15].lstrip().lower().startswith(b"<!doctype html") \
            or raw[:6].lower() == b"<html>":
        raise PermissionError(
            "シートを取得できませんでした。スプレッドシートの共有設定を"
            "「リンクを知っている全員が閲覧可」にしてください。"
        )

    text = raw.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return _build_history_from_rows(rows)


def load_history_from_file(path_or_bytes) -> dict[str, HistoryRecord]:
    """
    ローカルの履歴ファイル(.xlsx / .csv)から履歴を取得する(方式B用)。
    列構成: A型式 B品名 C発注先 D単価 E備考
    """
    if isinstance(path_or_bytes, str):
        is_csv = path_or_bytes.lower().endswith(".csv")
        data = None
    else:
        # bytes/file-like。先頭で判別が難しいので呼び出し側がxlsx前提なら xlsx として扱う。
        is_csv = False
        data = path_or_bytes if isinstance(path_or_bytes, bytes) else path_or_bytes.read()

    if is_csv:
        import csv
        with open(path_or_bytes, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            rows = list(csv.reader(f))
        return _build_history_from_rows(rows)

    # xlsx
    src = io.BytesIO(data) if data is not None else path_or_bytes
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rows = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    return _build_history_from_rows(rows)


# ----------------------------------------------------------------------------
# Excel 処理
# ----------------------------------------------------------------------------
def _detect_columns(ws) -> dict[str, int]:
    """1行目のヘッダーから列番号を推定する。見つからない項目は既定位置を使う。"""
    mapping = {
        "model": COL_MODEL, "maker": COL_MAKER, "qty": COL_QTY,
        "price": COL_PRICE, "vendor": COL_VENDOR, "total": COL_TOTAL,
        "spec_url": COL_SPECURL,
    }
    header = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            continue
        text = _norm_header(val)
        for field_key, aliases in HEADER_ALIASES.items():
            if any(_norm_header(a) in text for a in aliases):
                header.setdefault(field_key, col)
    mapping.update(header)
    return mapping


def build_spec_url(model: str, maker: str = "", engine: str = "google") -> str:
    """型式(必要ならメーカー名併記)で仕様を調べる検索URLを生成する。"""
    import urllib.parse
    model = (model or "").strip()
    if not model:
        return ""
    if engine == "misumi":
        return "https://jp.misumi-ec.com/vona2/result/?Keyword=" + urllib.parse.quote(model)
    if engine == "monotaro":
        return "https://www.monotaro.com/s/?q=" + urllib.parse.quote(model)
    # 既定: Google検索(メーカー名があれば併記して精度向上)
    q = f"{model} {maker}".strip() if maker else model
    return "https://www.google.com/search?q=" + urllib.parse.quote(q)


def process_workbook(
    source,
    history: dict[str, HistoryRecord],
    *,
    maker_map: Optional[dict[str, str]] = None,
    fill_total: bool = True,
    spec_engine: str = "google",
) -> tuple[bytes, ProcessResult]:
    """
    新規発注リスト(Excel)を読み込み、型式照合で補完して bytes を返す。

    列構成: A型式 B メーカー C数量 D単価 E品名 F合計金額 G仕様URL

    Parameters
    ----------
    source : str | bytes | file-like
        入力 .xlsx のパス、bytes、またはファイルオブジェクト。
    history : dict
        型式→HistoryRecord(単価・品名)の辞書。
    maker_map : dict[str,str] | None
        型式→メーカー名の対応表。B列メーカーの補完に使う。
    fill_total : bool
        True なら合計金額 F列 = 単価×数量 を計算して埋める。
    spec_engine : str
        仕様URLの検索先 ("google" / "misumi" / "monotaro")。

    Returns
    -------
    (output_bytes, ProcessResult)
    """
    maker_map = maker_map or {}
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    in_wb = load_workbook(source, data_only=True)
    in_ws = in_wb.active
    cols = _detect_columns(in_ws)
    model_col = cols["model"]
    qty_col = cols["qty"]

    # 出力は必ず正しい列見出しで作り直す
    # (A型式 B メーカー C数量 D単価 E発注先 F合計金額 G仕様URL)
    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "新規発注リスト"
    headers = ["型式", "メーカー", "数量", "単価", "発注先", "合計金額", "仕様URL"]
    widths = [22, 14, 8, 12, 16, 12, 40]
    bold = Font(bold=True)
    head_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = head_fill
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    result = ProcessResult()
    out_r = 1
    for r in range(2, in_ws.max_row + 1):
        model = _normalize_model(in_ws.cell(row=r, column=model_col).value)
        if not model:
            continue
        out_r += 1
        result.total_rows += 1

        qty_raw = in_ws.cell(row=r, column=qty_col).value
        qty = _parse_price(qty_raw)

        # 履歴から 単価・発注先 を補完
        rec = history.get(model)
        if rec is not None:
            price = rec.price
            vendor = rec.vendor or ""
            result.matched += 1
        else:
            price = None
            vendor = ""
            result.new_items += 1
            result.new_models.append(model)

        maker = maker_map.get(model, "")  # メーカー(対応表より)
        total = (price * qty) if (fill_total and price is not None and qty is not None) else None
        url = build_spec_url(model, maker, spec_engine)  # 仕様URL(検索リンク)

        ws.cell(row=out_r, column=COL_MODEL, value=model)
        ws.cell(row=out_r, column=COL_MAKER, value=maker or None)
        ws.cell(row=out_r, column=COL_QTY, value=qty_raw)
        ws.cell(row=out_r, column=COL_PRICE, value=price)
        ws.cell(row=out_r, column=COL_VENDOR, value=vendor or None)
        ws.cell(row=out_r, column=COL_TOTAL, value=total)
        ws.cell(row=out_r, column=COL_SPECURL, value=url)

        result.preview.append({
            "型式": model,
            "メーカー": maker,
            "数量": qty_raw,
            "単価": price,
            "発注先": vendor,
            "合計金額": total,
            "仕様URL": url,
        })

    out = io.BytesIO()
    out_wb.save(out)
    return out.getvalue(), result


# ----------------------------------------------------------------------------
# 雛形(テンプレート)生成
# ----------------------------------------------------------------------------
def build_template() -> bytes:
    """新規発注リストの空テンプレート(.xlsx)を生成して bytes で返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "新規発注リスト"
    headers = ["型式", "メーカー", "数量", "単価", "発注先", "合計金額", "仕様URL"]
    widths = [22, 14, 8, 12, 16, 12, 40]
    bold = Font(bold=True)
    head_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = head_fill
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    # 入力例を1行入れておく(型式と数量だけ入れて使う)
    ws.cell(row=2, column=COL_MODEL, value="(ここに型式)")
    ws.cell(row=2, column=COL_QTY, value=1)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------------------------------------------------------
# 型式→メーカー 対応表の読み込み
# ----------------------------------------------------------------------------
def build_maker_template() -> bytes:
    """型式→メーカー対応表の空テンプレート(.xlsx)を生成する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "型式メーカー対応表"
    bold = Font(bold=True)
    head_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for i, h in enumerate(["型式", "メーカー"], start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = head_fill
        ws.column_dimensions[get_column_letter(i)].width = 24
    ws.cell(row=2, column=1, value="(型式の例) F730")
    ws.cell(row=2, column=2, value="(メーカーの例) ユニパルス")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def load_maker_map_from_rows(rows) -> dict[str, str]:
    """行リストから 型式→メーカー の辞書を作る。
    見出しがあれば「型式」「メーカー」列を名前で判定、無ければ A=型式 B=メーカー。"""
    out: dict[str, str] = {}
    if not rows:
        return out
    idx = _find_header_indices(rows[0])
    if "model" in idx and "maker" in idx:
        mcol, kcol = idx["model"], idx["maker"]
        data = rows[1:]
    else:
        mcol, kcol = 0, 1
        data = rows
    for row in data:
        model = _normalize_model(row[mcol] if mcol < len(row) else "")
        if not model:
            continue
        maker = str(row[kcol]).strip() if kcol < len(row) and row[kcol] is not None else ""
        out[model] = maker
    return out


def load_maker_map_from_file(path_or_bytes) -> dict[str, str]:
    """型式→メーカー対応表(.xlsx / .csv)を読み込んで辞書を返す。"""
    if isinstance(path_or_bytes, str) and path_or_bytes.lower().endswith(".csv"):
        import csv
        with open(path_or_bytes, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            return load_maker_map_from_rows(list(csv.reader(f)))
    if isinstance(path_or_bytes, bytes):
        data = path_or_bytes
    elif hasattr(path_or_bytes, "read"):
        data = path_or_bytes.read()
    else:
        data = None
    src = io.BytesIO(data) if data is not None else path_or_bytes
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rows = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    return load_maker_map_from_rows(rows)
